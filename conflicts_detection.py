# -*- coding: utf-8 -*-
#%%
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors, Font, PatternFill
'''
# python36
import tkinter as tk
import tkinter.filedialog as filedialog
from tkinter import ttk, StringVar, Toplevel
'''
# python27
import Tkinter as tk 
import ttk
import tkFileDialog as filedialog
from Tkinter import StringVar, Toplevel

def conflict_check(new_word, word_dicts):
    check_result = []
    for name, word_dict in word_dicts.items():
        result = [key + 2 for (key, value) in word_dict.items() if (new_word !="empty" and new_word in value.split(","))]
        #print(name, word_dict, result)
        if (len(result)>0):
            check_result.append(name + ":" + str(result))
    if(len(check_result)>0):
        return new_word.encode('utf-8') + "-" + ",".join(check_result)
    else:
        return ''


#%%
'''
names = sys.argv
file_name = names[0]
sheet_name = names[1]
file_name = '游戏综合行业概念关系表V56.xls'
sheet_name = '概念'
data = pd.read_excel(file_name, sheet_name=sheet_name) \
    .iloc[:,[0, 1, 2, 3, 4, 5, 6, 7]].fillna("empty")
biaoshi_dict = data.iloc[:,0].to_dict()
gainian_dict = data.iloc[:,1].to_dict()
huozi_dict = data.iloc[:,2].to_dict()
dict_names = {
    "标识": biaoshi_dict,
    "概念": gainian_dict, 
    "或子": huozi_dict
}
tongyi_result = data.iloc[:,3].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split(",")  if conflict_check(w, dict_names) != '']))
jinyi_result = data.iloc[:,4].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split(",")  if conflict_check(w, dict_names) != '']))
fanshibie_result = data.iloc[:,5].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split("+")  if conflict_check(w, dict_names) != '']))
data["同义检测"] = tongyi_result
data["近义检测"] = jinyi_result
data["泛识别检测"] = fanshibie_result

cols = data.columns[:8]
columns_to_save = cols.insert(4, "同义检测").insert(6, "近义检测").insert(8, "泛识别检测")
data.to_excel(file_name.split(".")[0] + "_conflicts_check.xlsx", columns=columns_to_save)

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(data, index=False, header=True):
    ws.append(r)

for col in ws.iter_cols(min_col=9, max_col=11):
    for cell in col:
        if cell.value != '':
            cell.fill = PatternFill(fill_type='solid', fgColor=colors.RED)

wb.save(file_name.split(".")[0] + "_conflicts_check.xlsx")
'''

def check_output():
    data = pd.read_excel(file_name.get(), sheetname=1) \
        .iloc[:,[0, 1, 2, 3, 4, 5, 6, 7]].fillna("empty")
    biaoshi_dict = data.iloc[:,0].to_dict()
    gainian_dict = data.iloc[:,1].to_dict()
    huozi_dict = data.iloc[:,2].to_dict()
    dict_names = {
        "标识": biaoshi_dict,
        "概念": gainian_dict, 
        "或子": huozi_dict
    }
    tongyi_result = data.iloc[:,3].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split(",")  if conflict_check(w, dict_names) != '']))
    jinyi_result = data.iloc[:,4].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split(",")  if conflict_check(w, dict_names) != '']))
    fanshibie_result = data.iloc[:,5].apply(lambda x: "\r\n".join([conflict_check(w, dict_names) for w in x.split("+")  if conflict_check(w, dict_names) != '']))
    data["同义检测"] = tongyi_result
    data["近义检测"] = jinyi_result
    data["泛识别检测"] = fanshibie_result
    '''
    cols = data.columns[:8]
    columns_to_save = cols.insert(4, "同义检测").insert(6, "近义检测").insert(8, "泛识别检测")
    data.to_excel(file_name.split(".")[0] + "_conflicts_check.xlsx", columns=columns_to_save)
    '''
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    for col in ws.iter_cols(min_col=9, max_col=11):
        for cell in col:
            if cell.value != '':
                cell.fill = PatternFill(fill_type='solid', fgColor=colors.RED)

    wb.save(file_name.get().split(".")[0] + "_conflicts_check.xlsx")
    
    
# conflict_check("页游数据", dict_names)
#%%
def get_screen_size(window):  
    return window.winfo_screenwidth(),window.winfo_screenheight()  
  
def get_window_size(window):  
    return window.winfo_reqwidth(),window.winfo_reqheight()  
  
def center_window(root, width, height):  
    screenwidth = root.winfo_screenwidth()  
    screenheight = root.winfo_screenheight()  
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width)/2, (screenheight - height)/2)  
    #print(size)  
    root.geometry(size)

def no_click():
    popup = Toplevel()
    popup.resizable(False,False)
    ttk.Label(popup, text='请选择文件!').pack(pady=25)
    center_window(popup, 300, 100)
    popup.after(2500,lambda:popup.destroy())

def choose_file():
    file_name_1 = filedialog.askopenfilename()
    if file_name_1 != '':
        path.set(file_name_1)
        file_name.set(file_name_1)
    elif path.get() == '':
        no_click()
    else:
        return

root = tk.Tk(className="冲突检测")
sheet_name = 1
center_window(root, 600, 100)
file_select_label = ttk.Label(root, text="待检测文档：")
file_select_label.grid(column=1, row=0)
path = StringVar()
file_name = StringVar()
duibiao_file_path = ttk.Entry(root, textvariable=path, width=50, state=['readonly'])
duibiao_file_path.grid(column=2, row=0)
choose_file_action = ttk.Button(root,text="浏览", width=10, command=choose_file)
choose_file_action.grid(column=3, row=0)
check_action = ttk.Button(root,text="检测", width=10, command=check_output)
check_action.grid(column=2, row=1)
for child in root.winfo_children():  
    child.grid(padx=5, pady=5) 
root.mainloop()


